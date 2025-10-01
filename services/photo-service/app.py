import os
import io
import json
import zipfile
import imghdr
import uuid as _uuid
from math import radians, sin, cos, sqrt, atan2
from pathlib import Path
from datetime import datetime, date, timedelta
import random

from flask import Flask, request, jsonify, send_from_directory, abort
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import openpyxl
except Exception:
    openpyxl = None

import psycopg  # psycopg v3 (единый драйвер для всего файла)

# -----------------------------------------------------------------------------
# Конфиг
# -----------------------------------------------------------------------------
DB_DSN = (
    f"dbname={os.getenv('POSTGRES_DB', 'geolocate_db')} "
    f"user={os.getenv('POSTGRES_USER', 'geolocate_user')} "
    f"password={os.getenv('POSTGRES_PASSWORD', 'StrongPass123!')} "
    f"host={os.getenv('POSTGRES_HOST', 'db')} "
    f"port={os.getenv('POSTGRES_PORT', '5432')}"
)

# Папки
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "uploads")).resolve()
INSTALL_DIR = Path(os.getenv("INSTALL_DIR", "/app/install")).resolve()
MODEL_DIR = Path(os.getenv("MODEL_DIR", "/app/model")).resolve()
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MODEL_DIR.mkdir(parents=True, exist_ok=True)

# Ограничение размера тела запроса (для Flask)
MAX_CONTENT_LENGTH = int(os.getenv("MAX_UPLOAD_MB", "2048")) * 1024 * 1024  # дефолт 2 ГБ

# Центр для имитации «расчёта»
CENTER_LAT = float(os.getenv("CENTER_LAT", 55.804111))
CENTER_LON = float(os.getenv("CENTER_LON", 37.749822))

# Ограничения ZIP
ZIP_MAX_FILES = int(os.getenv("ZIP_MAX_FILES", "500"))
ZIP_MAX_UNCOMPRESSED_MB = int(os.getenv("ZIP_MAX_UNCOMPRESSED_MB", "4096"))  # 4 ГБ
ALLOWED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tif", ".tiff"}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

# -----------------------------------------------------------------------------
# БД: соединение + схема с has_coords (STORED)
# -----------------------------------------------------------------------------
def get_conn():
    return psycopg.connect(DB_DSN)

def ensure_schema():
    with get_conn() as conn, conn.cursor() as cur:
        # photos
        cur.execute("""
            CREATE TABLE IF NOT EXISTS photos (
                id         BIGSERIAL PRIMARY KEY,
                created    TIMESTAMPTZ NOT NULL DEFAULT now(),

                name       VARCHAR(512) NOT NULL,
                uuid       UUID UNIQUE NOT NULL,

                width      INTEGER,
                height     INTEGER,

                exif_lat   DOUBLE PRECISION,
                exif_lon   DOUBLE PRECISION,

                type       VARCHAR(64),
                subtype    VARCHAR(64),

                shot_lat   DOUBLE PRECISION,
                shot_lon   DOUBLE PRECISION
            );
        """)
        # detected_objects
        cur.execute("""
            CREATE TABLE IF NOT EXISTS detected_objects (
                id           BIGSERIAL PRIMARY KEY,
                photo_id     BIGINT NOT NULL REFERENCES photos(id) ON DELETE CASCADE,
                label        VARCHAR(64) NOT NULL DEFAULT 'object',
                confidence   DOUBLE PRECISION DEFAULT 0.0,
                x1           INTEGER,
                y1           INTEGER,
                x2           INTEGER,
                y2           INTEGER,
                latitude     DOUBLE PRECISION,
                longitude    DOUBLE PRECISION,
                created      TIMESTAMPTZ NOT NULL DEFAULT now()
            );
        """)
        # history
        cur.execute("""
            CREATE TABLE IF NOT EXISTS history (
                id       BIGSERIAL PRIMARY KEY,
                created  TIMESTAMPTZ NOT NULL DEFAULT now(),
                event    VARCHAR(64) NOT NULL,
                payload  JSONB
            );
        """)

        # Генерируемый столбец has_coords (если нет)
        # Postgres 12+ поддерживает GENERATED ... STORED
        try:
            cur.execute("""
                ALTER TABLE photos
                ADD COLUMN IF NOT EXISTS has_coords boolean
                GENERATED ALWAYS AS (shot_lat IS NOT NULL AND shot_lon IS NOT NULL) STORED;
            """)
        except Exception:
            # На случай старой версии PG — создадим обычный столбец и будем обновлять триггером (упростим: просто столбец).
            cur.execute("DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='photos' AND column_name='has_coords') THEN ALTER TABLE photos ADD COLUMN has_coords boolean; END IF; END $$;")
            # Одноразово проставим текущие значения
            cur.execute("UPDATE photos SET has_coords = (shot_lat IS NOT NULL AND shot_lon IS NOT NULL) WHERE has_coords IS NULL;")

        # Индексы
        cur.execute("CREATE INDEX IF NOT EXISTS idx_photos_uuid          ON photos (uuid);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_photos_created       ON photos (created DESC);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_photos_has_created   ON photos (has_coords, created DESC);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_photos_shot          ON photos (shot_lat, shot_lon);")

        cur.execute("CREATE INDEX IF NOT EXISTS idx_detected_photo_created ON detected_objects (photo_id, created DESC);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_detected_label         ON detected_objects (label);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_detected_geo           ON detected_objects (latitude, longitude);")

ensure_schema()

def _insert_history(event, payload):
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("INSERT INTO history(event, payload) VALUES (%s, %s)", (event, json.dumps(payload)))
    except Exception:
        pass

# -----------------------------------------------------------------------------
# Утилиты
# -----------------------------------------------------------------------------
def _img_size_from_bytes(buf: bytes):
    if not Image:
        return None, None
    try:
        with Image.open(io.BytesIO(buf)) as im:
            return im.size
    except Exception:
        return None, None

def _img_size_from_path(path: Path):
    if not Image:
        return None, None
    try:
        with Image.open(path) as im:
            return im.size
    except Exception:
        return None, None

def _to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _parse_bool(s: str):
    if s is None:
        return None
    v = str(s).strip().lower()
    if v in ("true", "1", "yes", "y", "on", "with"):
        return True
    if v in ("false", "0", "no", "n", "off", "without"):
        return False
    if v in ("all", ""):
        return None
    return None

def _ymd(s: str):
    # ожидаем YYYY-MM-DD
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _ext(name: str) -> str:
    return os.path.splitext(name.lower())[-1]

def _safe_member_name(name: str) -> str:
    name = name.replace("\\", "/").lstrip("/")
    parts = [p for p in name.split("/") if p not in ("", ".", "..")]
    return "/".join(parts)

def _save_photo_record(saved_name, uid, width, height, ptype, subtype, shot_lat, shot_lon):
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            INSERT INTO photos (name, uuid, width, height, type, subtype, shot_lat, shot_lon)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (uuid) DO NOTHING
            RETURNING id
        """, (saved_name, uid, width, height, ptype, subtype, shot_lat, shot_lon))
        row = cur.fetchone()
        if row:
            return row[0]
        cur.execute("SELECT id FROM photos WHERE uuid=%s", (uid,))
        row = cur.fetchone()
        return row[0] if row else None

# -----------------------------------------------------------------------------
# Импорт из /app/install при первом старте (опционально)
# -----------------------------------------------------------------------------
def _read_excel_meta(xlsx_path: Path):
    meta = {}
    if not xlsx_path.exists():
        return meta
    # openpyxl
    if openpyxl:
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            headers = {}
            for j, cell in enumerate(ws[1], start=1):
                key = str(cell.value or "").strip().lower()
                headers[key] = j
            fn_col = headers.get("имя файла") or headers.get("filename") or headers.get("файл")
            cam_col = headers.get("camera") or headers.get("камера")
            lat_col = headers.get("latitude") or headers.get("широта")
            lon_col = headers.get("longitude") or headers.get("долгота")
            for i in range(2, ws.max_row + 1):
                name = ws.cell(i, fn_col).value if fn_col else None
                if not name:
                    continue
                camera = ws.cell(i, cam_col).value if cam_col else None
                lat = _to_float(ws.cell(i, lat_col).value) if lat_col else None
                lon = _to_float(ws.cell(i, lon_col).value) if lon_col else None
                meta[str(name).strip()] = {"camera": camera, "lat": lat, "lon": lon}
            return meta
        except Exception:
            pass
    # CSV fallback (;)
    try:
        import csv
        with open(xlsx_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                name = (row.get("Имя файла") or row.get("filename") or row.get("файл") or "").strip()
                if not name:
                    continue
                cam = row.get("camera") or row.get("камера")
                lat = _to_float(row.get("latitude") or row.get("широта"))
                lon = _to_float(row.get("longitude") or row.get("долгота"))
                meta[name] = {"camera": cam, "lat": lat, "lon": lon}
    except Exception:
        pass
    return meta

def _read_json_results(json_path: Path):
    mapping = {}
    if not json_path.exists():
        return mapping
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
        results = data.get("results") or []
        for rec in results:
            fname_guess = None
            if isinstance(rec.get("id"), str):
                fname_guess = rec["id"] + ".jpg"
            if not fname_guess and isinstance(rec.get("image"), str):
                tail = rec["image"].rstrip("/").split("/")[-1]
                if tail:
                    fname_guess = tail + ".jpg"
            shot_lat = _to_float(rec.get("latitude"))
            shot_lon = _to_float(rec.get("longitude"))
            issues = rec.get("issues") or []
            if fname_guess:
                mapping[fname_guess] = {"lat": shot_lat, "lon": shot_lon, "issues": issues}
    except Exception:
        pass
    return mapping

def _import_from_install():
    # Только если БД пустая
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM photos")
        count = cur.fetchone()[0]
    if count > 0 or not INSTALL_DIR.exists():
        return

    # model.zip → MODEL_DIR
    model_zip = INSTALL_DIR / "model.zip"
    if model_zip.exists():
        try:
            with zipfile.ZipFile(model_zip, "r") as z:
                z.extractall(MODEL_DIR)
        except Exception:
            pass

    imported = 0
    created_objects = 0

    for arch in INSTALL_DIR.glob("*.zip"):
        stem = arch.stem
        # JSON поблизости
        json_path = INSTALL_DIR / f"{stem}.json"
        if not json_path.exists():
            cand = list(INSTALL_DIR.glob(f"{stem}*.json"))
            json_path = cand[0] if cand else json_path
        # Excel/CSV поблизости
        excel_path = None
        for ext in ("xlsx", "xls", "csv"):
            p = INSTALL_DIR / f"{stem}.{ext}"
            if p.exists():
                excel_path = p
                break

        json_map = _read_json_results(json_path) if json_path and json_path.exists() else {}
        excel_map = _read_excel_meta(excel_path) if excel_path else {}

        try:
            with zipfile.ZipFile(arch, "r") as z:
                for info in z.infolist():
                    if info.is_dir():
                        continue
                    name = Path(info.filename).name
                    if not name.lower().endswith(tuple(ALLOWED_EXTS)):
                        continue
                    raw = z.read(info)
                    kind = imghdr.what(None, h=raw) or "jpeg"

                    saved_name = name  # стараемся сохранить оригинальное имя
                    dest = UPLOAD_DIR / saved_name
                    if not dest.exists():
                        dest.write_bytes(raw)

                    width, height = _img_size_from_bytes(raw)

                    em = excel_map.get(name) or {}
                    jm = json_map.get(name) or {}
                    shot_lat = em.get("lat") if em.get("lat") is not None else jm.get("lat")
                    shot_lon = em.get("lon") if em.get("lon") is not None else jm.get("lon")
                    ptype = "unknown"
                    subtype = "unknown"

                    try:
                        uid_val = str(_uuid.UUID(Path(name).stem))
                    except Exception:
                        uid_val = str(_uuid.uuid4())

                    pid = _save_photo_record(saved_name, uid_val, width, height, ptype, subtype, shot_lat, shot_lon)
                    if pid:
                        imported += 1
                        issues = (jm.get("issues") or [])
                        with get_conn() as conn, conn.cursor() as cur:
                            for iss in issues:
                                bbox = iss.get("bbox") or {}
                                x = int(bbox.get("x", 10))
                                y = int(bbox.get("y", 10))
                                w = int(bbox.get("w", 120))
                                h = int(bbox.get("h", 120))
                                x1, y1, x2, y2 = x, y, x + w, y + h
                                conf = float(iss.get("score", 0.8))
                                lat = _to_float(iss.get("latitude")) or shot_lat
                                lon = _to_float(iss.get("longitude")) or shot_lon
                                cur.execute("""
                                    INSERT INTO detected_objects (photo_id,label,confidence,x1,y1,x2,y2,latitude,longitude)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                """, (pid, iss.get("label") or "object", conf, x1,y1,x2,y2, lat, lon))
                                created_objects += 1
        except Exception as e:
            _insert_history("init_import_archive_error", {"archive": arch.name, "error": str(e)})

    _insert_history("init_import", {"imported_photos": imported, "created_objects": created_objects})

try:
    _import_from_install()
except Exception as e:
    _insert_history("init_import_error", {"error": str(e)})

# -----------------------------------------------------------------------------
# Маршруты
# -----------------------------------------------------------------------------
@app.errorhandler(RequestEntityTooLarge)
def too_large(e):
    return {"error": f"file too large (> {app.config['MAX_CONTENT_LENGTH']//1024//1024} MB)"}, 413

@app.get("/healthz")
def healthz():
    return {"status": "ok", "service": "photo"}

# -- LIST с серверными фильтрами has_coords + date_from/date_to
@app.get("/photos")
def photos_list():
    # Пагинация
    try:
        limit = int(request.args.get("limit", "50"))
    except Exception:
        limit = 50
    try:
        offset = int(request.args.get("offset", "0"))
    except Exception:
        offset = 0
    limit = max(1, min(limit, 500))
    offset = max(0, offset)

    # Фильтры
    has_coords = _parse_bool(request.args.get("has_coords"))  # True/False/None
    date_from = _ymd(request.args.get("date_from") or "")
    date_to   = _ymd(request.args.get("date_to") or "")
    # включительно: [from, to 23:59:59.999]
    if date_to:
        date_to_excl = datetime.combine(date_to + timedelta(days=1), datetime.min.time())
    else:
        date_to_excl = None

    where = []
    params = []

    if has_coords is True:
        where.append("has_coords = TRUE")
    elif has_coords is False:
        where.append("has_coords = FALSE")

    if date_from:
        where.append("created >= %s")
        params.append(datetime.combine(date_from, datetime.min.time()))
    if date_to_excl:
        where.append("created < %s")
        params.append(date_to_excl)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    with get_conn() as conn, conn.cursor() as cur:
        # total по ОТРЕЗАННОМУ набору (с учётом фильтров) — так корректно работает пагинация на фронте
        cur.execute(f"SELECT COUNT(*) FROM photos {where_sql}", params)
        total = cur.fetchone()[0]

        cur.execute(f"""
            SELECT id, uuid::text, name, width, height, created, type, subtype, shot_lat, shot_lon, has_coords
            FROM photos
            {where_sql}
            ORDER BY created DESC, id DESC
            LIMIT %s OFFSET %s
        """, [*params, limit, offset])
        rows = cur.fetchall()

    return {
        "total": total,
        "photos": [
            {
                "id": r[0], "uuid": r[1], "name": r[2],
                "width": r[3], "height": r[4],
                "created": r[5].isoformat() if r[5] else None,
                "type": r[6], "subtype": r[7],
                "shot_lat": r[8], "shot_lon": r[9],
                "has_coords": r[10],
            }
            for r in rows
        ]
    }

@app.get("/photos/<uuid_str>")
def photo_file(uuid_str: str):
    try:
        _ = _uuid.UUID(uuid_str)
    except Exception:
        abort(404)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT name FROM photos WHERE uuid=%s", (uuid_str,))
        row = cur.fetchone()
        if not row:
            abort(404)
        fname = row[0]
    return send_from_directory(str(UPLOAD_DIR), fname, as_attachment=False)

# --- Одиночная загрузка
@app.post("/upload")
def upload():
    """
    multipart/form-data:
      image: файл изображения (обязательно)
      type, subtype: строки
      meta: json-файл с {"lat": ..., "lon": ...} (опц.)
      shot_lat / shot_lon: числа (если meta нет)
    """
    if "image" not in request.files:
        return {"error": "no file field 'image'"}, 400
    f = request.files["image"]
    if f.filename == "":
        return {"error": "empty filename"}, 400

    ptype = (request.form.get("type") or "unknown")[:64]
    subtype = (request.form.get("subtype") or "unknown")[:64]

    shot_lat = request.form.get("shot_lat")
    shot_lon = request.form.get("shot_lon")
    if "meta" in request.files and request.files["meta"].filename:
        try:
            meta_data = json.load(io.TextIOWrapper(request.files["meta"].stream, encoding="utf-8"))
            shot_lat = meta_data.get("lat", shot_lat)
            shot_lon = meta_data.get("lon", shot_lon)
        except Exception:
            pass

    shot_lat = _to_float(shot_lat) if shot_lat is not None else None
    shot_lon = _to_float(shot_lon) if shot_lon is not None else None

    raw = f.read()
    if not raw:
        return {"error": "empty file"}, 400
    kind = imghdr.what(None, h=raw)
    if kind not in {"jpeg", "png", "gif", "tiff", "bmp", "webp"}:
        return {"error": "unsupported image"}, 415

    orig_name = secure_filename(f.filename)
    saved_name = orig_name
    dest = UPLOAD_DIR / saved_name
    i = 1
    while dest.exists():
        stem = Path(orig_name).stem
        ext = Path(orig_name).suffix or (".jpg" if kind == "jpeg" else f".{kind}")
        saved_name = f"{stem}_{i}{ext}"
        dest = UPLOAD_DIR / saved_name
        i += 1
    dest.write_bytes(raw)

    width, height = _img_size_from_bytes(raw)

    try:
        uid_val = str(_uuid.UUID(Path(saved_name).stem))
    except Exception:
        uid_val = str(_uuid.uuid4())

    pid = _save_photo_record(saved_name, uid_val, width, height, ptype, subtype, shot_lat, shot_lon)

    return {
        "photo": {
            "id": pid, "uuid": uid_val, "name": saved_name,
            "width": width, "height": height,
            "type": ptype, "subtype": subtype,
            "shot_lat": shot_lat, "shot_lon": shot_lon
        }
    }, 201

# --- Массовый импорт ZIP (sidecar/manifest/defaults)
def _merge_meta(defaults, manifest_entry, sidecar):
    out = dict(defaults or {})
    if manifest_entry:
        out.update({k: v for k, v in (manifest_entry.items()) if v not in (None, "")})
    if sidecar:
        out.update({k: v for k, v in (sidecar.items()) if v not in (None, "")})
    # normalize
    for k in ("lat", "lon"):
        if k in out:
            try:
                out[k] = float(out[k])
            except Exception:
                out.pop(k, None)
    # маппинг shot_lat/shot_lon
    if "shot_lat" in out and "lat" not in out:
        out["lat"] = _to_float(out["shot_lat"])
    if "shot_lon" in out and "lon" not in out:
        out["lon"] = _to_float(out["shot_lon"])
    return out

@app.post("/upload_zip")
def upload_zip():
    f = request.files.get("archive") or request.files.get("zip")
    if not f:
        return jsonify(error="no file field 'archive' or 'zip'"), 400

    defaults = {}
    if request.form.get("type"):     defaults["type"] = request.form["type"]
    if request.form.get("subtype"):  defaults["subtype"] = request.form["subtype"]
    if request.form.get("shot_lat"): defaults["lat"] = request.form["shot_lat"]
    if request.form.get("shot_lon"): defaults["lon"] = request.form["shot_lon"]

    data = f.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception as e:
        return jsonify(error=f"bad zip: {e}"), 400

    names = [n for n in zf.namelist() if not n.endswith("/")]
    if len(names) > ZIP_MAX_FILES:
        return jsonify(error=f"too many files in zip (> {ZIP_MAX_FILES})"), 400

    total_uncompressed = sum(zi.file_size for zi in zf.infolist())
    if total_uncompressed > ZIP_MAX_UNCOMPRESSED_MB * 1024 * 1024:
        return jsonify(error=f"uncompressed size exceeds {ZIP_MAX_UNCOMPRESSED_MB} MB"), 400

    # manifest.json (опционально)
    manifest_map = {}
    if "manifest.json" in names:
        try:
            raw = zf.read("manifest.json")
            m = json.loads(raw.decode("utf-8", "ignore"))
            for it in (m.get("items") or []):
                fname = _safe_member_name(it.get("file") or "")
                if fname:
                    manifest_map[fname] = it
        except Exception:
            manifest_map = {}

    results, errors = [], []

    with get_conn() as conn, conn.cursor() as cur:
        try:
            for member in names:
                safe_name = _safe_member_name(member)
                ext = _ext(safe_name)
                if ext not in ALLOWED_EXTS:
                    if ext == ".json":
                        continue  # sidecar — ок, просто пропустим
                    errors.append({"file": member, "error": "unsupported extension"})
                    continue

                try:
                    # sidecar (file.jpg -> file.json)
                    base, _ = os.path.splitext(safe_name)
                    sidecar_json = None
                    if f"{base}.json" in names:
                        try:
                            sidecar_raw = zf.read(f"{base}.json")
                            sidecar_json = json.loads(sidecar_raw.decode("utf-8", "ignore"))
                        except Exception:
                            sidecar_json = None

                    manifest_entry = manifest_map.get(safe_name)
                    meta = _merge_meta(defaults, manifest_entry, sidecar_json)

                    img_bytes = zf.read(member)
                    w, h = _img_size_from_bytes(img_bytes)

                    file_uuid = str(_uuid.uuid4())
                    stored_name = f"{file_uuid}{ext}"
                    out_path = UPLOAD_DIR / stored_name
                    out_path.write_bytes(img_bytes)

                    cur.execute("""
                        INSERT INTO photos (name, uuid, width, height, type, subtype, shot_lat, shot_lon)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id, created
                    """, (
                        stored_name, file_uuid, w, h,
                        meta.get("type"), meta.get("subtype"),
                        meta.get("lat"), meta.get("lon"),
                    ))
                    pid, created = cur.fetchone()

                    results.append({
                        "file": member,
                        "id": pid,
                        "uuid": file_uuid,
                        "name": stored_name,
                        "width": w,
                        "height": h,
                        "lat": meta.get("lat"),
                        "lon": meta.get("lon"),
                        "type": meta.get("type"),
                        "subtype": meta.get("subtype"),
                        "created": created.isoformat() if isinstance(created, datetime) else str(created),
                    })
                except Exception as e:
                    errors.append({"file": member, "error": str(e)})

            conn.commit()
        except Exception as e:
            conn.rollback()
            return jsonify(error=str(e)), 500

    return jsonify({
        "ok": True,
        "imported": len(results),
        "skipped": len(errors),
        "results": results,
        "errors": errors
    })

# --- Поиск ближайших фото к точке (alias: /search_coords)
def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dl/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

@app.get("/search_coords")
@app.get("/search")  # совместимость
def search_by_coords():
    try:
        lat = float(request.args.get("lat"))
        lon = float(request.args.get("lon"))
    except Exception:
        return {"error": "lat/lon required"}, 400

    try:
        limit = int(request.args.get("limit", "12"))
    except Exception:
        limit = 12
    limit = max(1, min(limit, 200))

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, uuid::text, name, shot_lat, shot_lon
            FROM photos
            WHERE shot_lat IS NOT NULL AND shot_lon IS NOT NULL
        """)
        rows = cur.fetchall()

    items = []
    for pid, uid, name, slat, slon in rows:
        try:
            dist = haversine_m(lat, lon, float(slat), float(slon))
        except Exception:
            continue
        items.append({
            "id": pid, "uuid": uid, "name": name,
            "dist_m": dist, "shot_lat": slat, "shot_lon": slon
        })

    items.sort(key=lambda x: x["dist_m"])
    res = items[:limit]

    _insert_history("search_knn", {"lat": lat, "lon": lon, "limit": limit, "returned": len(res)})
    return {"results": res}

# --- Имитация «расчёта координат»
@app.post("/calc_for_photo")
def calc_for_photo():
    data = request.get_json(silent=True) or {}
    try:
        photo_id = int(data.get("photo_id"))
    except Exception:
        return {"error": "photo_id required"}, 400

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id FROM photos WHERE id=%s", (photo_id,))
        row = cur.fetchone()
        if not row:
            return {"error": "photo not found"}, 404

        def jitter(lat, lon, meters=50.0):
            dlat = (random.uniform(-meters, meters)) / 111320.0
            dlng = (random.uniform(-meters, meters)) / (111320.0 * max(cos(radians(lat)), 1e-6))
            return lat + dlat, lon + dlng

        dets = []
        for bx in [(10,10,120,120), (140,20,260,180)]:
            lat, lon = jitter(CENTER_LAT, CENTER_LON, 50.0)
            dets.append({"label": "house", "confidence": 0.89, "bbox": bx, "lat": lat, "lon": lon})

        for d in dets:
            x1, y1, x2, y2 = d["bbox"]
            cur.execute("""
                INSERT INTO detected_objects
                    (photo_id,label,confidence,x1,y1,x2,y2,latitude,longitude)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (photo_id, d["label"], d["confidence"], x1, y1, x2, y2, d["lat"], d["lon"]))

    _insert_history("calc", {"photo_id": photo_id, "created": len(dets)})
    return {"message": f"Проанализирована 1 фото, обнаружено {len(dets)} дома. Уверенность — 89%."}, 200

# --- Метаданные по фото
@app.get("/photo_meta")
def photo_meta():
    try:
        photo_id = int(request.args.get("id", "0"))
    except Exception:
        return {"error": "id required"}, 400
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, uuid::text, name, width, height, shot_lat, shot_lon, created
            FROM photos WHERE id=%s
        """, (photo_id,))
        row = cur.fetchone()
        if not row:
            return {"error": "photo not found"}, 404
        return {
            "photo": {
                "id": row[0], "uuid": row[1], "name": row[2],
                "width": row[3], "height": row[4],
                "shot_lat": row[5], "shot_lon": row[6],
                "created": row[7].isoformat() if row[7] else None
            }
        }

# --- Вставка одной детекции
@app.post("/detect")
def detect_insert():
    data = request.get_json(silent=True) or {}
    try:
        photo_id = int(data.get("photo_id"))
    except Exception:
        return {"error": "photo_id required"}, 400

    label = (data.get("label") or "house")[:64]
    try:
        conf = float(data.get("confidence") or 0.0)
    except Exception:
        conf = 0.0

    x1 = data.get("x1"); y1 = data.get("y1"); x2 = data.get("x2"); y2 = data.get("y2")
    bbox = data.get("bbox")
    if isinstance(bbox, dict) and all(k in bbox for k in ("x","y","w","h")):
        x1 = int(bbox["x"]); y1 = int(bbox["y"])
        x2 = x1 + int(bbox["w"]); y2 = y1 + int(bbox["h"])

    try:
        x1 = int(x1); y1 = int(y1); x2 = int(x2); y2 = int(y2)
    except Exception:
        return {"error": "invalid bbox"}, 400

    lat = data.get("lat") or data.get("latitude")
    lon = data.get("lon") or data.get("longitude")
    try:
        lat = float(lat) if lat is not None else None
        lon = float(lon) if lon is not None else None
    except Exception:
        lat = lon = None

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            INSERT INTO detected_objects (photo_id,label,confidence,x1,y1,x2,y2,latitude,longitude)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (photo_id, label, conf, x1,y1,x2,y2, lat, lon))
        new_id = cur.fetchone()[0]
    return {"id": new_id}, 201

# --- Пакетная вставка детекций
@app.post("/detect_bulk")
def detect_bulk():
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    inserted = 0
    with get_conn() as conn, conn.cursor() as cur:
        for it in items:
            try:
                photo_id = int(it.get("photo_id"))
                label = (it.get("label") or "house")[:64]
                conf = float(it.get("confidence") or 0.0)
                bbox = it.get("bbox")
                x1 = it.get("x1"); y1 = it.get("y1"); x2 = it.get("x2"); y2 = it.get("y2")
                if isinstance(bbox, dict) and all(k in bbox for k in ("x","y","w","h")):
                    x1 = int(bbox["x"]); y1 = int(bbox["y"])
                    x2 = x1 + int(bbox["w"]); y2 = y1 + int(bbox["h"])
                x1 = int(x1); y1 = int(y1); x2 = int(x2); y2 = int(y2)
                lat = it.get("lat") or it.get("latitude")
                lon = it.get("lon") or it.get("longitude")
                lat = float(lat) if lat is not None else None
                lon = float(lon) if lon is not None else None
            except Exception:
                continue

            cur.execute("""
                INSERT INTO detected_objects (photo_id,label,confidence,x1,y1,x2,y2,latitude,longitude)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (photo_id, label, conf, x1,y1,x2,y2, lat, lon))
            inserted += 1
    return {"inserted": inserted}, 200

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")))

